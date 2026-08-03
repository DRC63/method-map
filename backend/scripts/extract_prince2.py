"""One-off: turn the PRINCE2 7 Cross Reference sheet into a clean, self-contained
JSON seed for the Method Map app. Entities are referenced by a composite
"type::name" key; relationships link an activity to a role/practice/approach/product
with a C/P/N (roles/practices/approaches) or I/O/U/A (products) code.
"""
import json
import os

import openpyxl

SRC = r"C:/Users/drcol/OneDrive/Documents/Methodologies/PRINCE2_MSP_Updated_Cross_Reference_Skeleton.xlsx"
OUT = r"C:/Users/drcol/OneDrive/Documents/Claude/claude-code/method-map/backend/app/seed_data/prince2-7.json"

# 1-indexed column ranges (see header row 2), classified by the row-1 group bands.
ROLE_COLS = range(4, 11)        # D-J
PRACTICE_COLS = range(12, 19)   # L-R
APPROACH_COLS = range(20, 29)   # T-AB
PRODUCT_BASELINE = range(30, 37)  # AD-AJ
PRODUCT_LOG = range(37, 43)       # AK-AP
PRODUCT_REPORT = range(43, 50)    # AQ-AW

ROLE_CODES = {"C", "P", "N"}
PRODUCT_CODES = {"I", "O", "U", "A"}

# Canonical PRINCE2 7 process-model placement (the timeline / swimlanes). Keyed
# by process code. level = swimlane, phase = where on the timeline it sits,
# sequence = left-to-right order, repeats = runs once per delivery stage.
LIFECYCLE = {
    "SU": {"lifecycle_level": "managing", "lifecycle_phase": "pre-project", "sequence": 1, "repeats": False},
    "DP": {"lifecycle_level": "directing", "lifecycle_phase": "throughout", "sequence": 2, "repeats": False},
    "IP": {"lifecycle_level": "managing", "lifecycle_phase": "initiation", "sequence": 3, "repeats": False},
    "CS": {"lifecycle_level": "managing", "lifecycle_phase": "delivery", "sequence": 4, "repeats": True},
    "MP": {"lifecycle_level": "delivering", "lifecycle_phase": "delivery", "sequence": 5, "repeats": True},
    "SB": {"lifecycle_level": "managing", "lifecycle_phase": "stage-boundary", "sequence": 6, "repeats": True},
    "CP": {"lifecycle_level": "managing", "lifecycle_phase": "final", "sequence": 7, "repeats": False},
}

wb = openpyxl.load_workbook(SRC, data_only=True)
ws = wb["PRINCE2 7 Cross Reference"]

# header labels from row 2
hdr = {}
for j, v in enumerate(next(ws.iter_rows(min_row=2, max_row=2, values_only=True)), start=1):
    hdr[j] = (v or "").strip() if isinstance(v, str) else v


def col_meta(j):
    """Return (entity_type, subgroup) for a data column, or (None, None)."""
    if j in ROLE_COLS:
        return "role", None
    if j in PRACTICE_COLS:
        return "practice", None
    if j in APPROACH_COLS:
        return "approach", None
    if j in PRODUCT_BASELINE:
        return "product", "baseline"
    if j in PRODUCT_LOG:
        return "product", "log"
    if j in PRODUCT_REPORT:
        return "product", "report"
    return None, None


entities = {}   # key "type::name" -> entity dict
relationships = []
order = {"process": 0, "activity": 0, "role": 0, "practice": 0, "approach": 0, "product": 0}


def add_entity(etype, name, **extra):
    key = f"{etype}::{name}"
    if key not in entities:
        order[etype] += 1
        entities[key] = {
            "type": etype,
            "name": name,
            "code": extra.get("code"),
            "subgroup": extra.get("subgroup"),
            "parent": extra.get("parent"),
            "confidence": extra.get("confidence", "confirmed"),
            "sort_order": order[etype],
            "lifecycle_level": extra.get("lifecycle_level"),
            "lifecycle_phase": extra.get("lifecycle_phase"),
            "sequence": extra.get("sequence"),
            "repeats": extra.get("repeats", False),
        }
    return key


# Pre-register the column-header entities (roles, practices, approaches, products)
# so they exist even if no activity happens to reference them.
for j in range(4, 50):
    etype, subgroup = col_meta(j)
    label = hdr.get(j)
    if etype and label:
        add_entity(etype, label, subgroup=subgroup, confidence="confirmed")

current_process = None
for row in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True):
    proc, act = row[0], row[1]
    proc = proc.strip() if isinstance(proc, str) else proc
    act = act.strip() if isinstance(act, str) else act

    if proc and not act:
        # process header row, e.g. "SU - Starting Up A Project"
        code, _, pname = proc.partition(" - ")
        code = code.strip()
        pname = pname.strip() or proc
        current_process = add_entity(
            "process", pname, code=code, confidence="confirmed", **LIFECYCLE.get(code, {})
        )
        continue

    if not act:
        continue

    # activity row (icons are a best-effort reconstruction -> indicative)
    akey = add_entity("activity", act, parent=current_process, confidence="indicative")

    for j, v in enumerate(row, start=1):
        if j < 3 or v in (None, ""):
            continue
        etype, subgroup = col_meta(j)
        if not etype:
            continue
        label = hdr.get(j)
        code = str(v).strip().upper()
        tkey = f"{etype}::{label}"
        relationships.append({
            "from": akey,
            "to": tkey,
            "code": code,
            "confidence": "indicative",
        })

# Manual corrections / additions not present in the source spreadsheet, applied
# after parsing so they survive regeneration. (from_key, to_key, code, confidence)
CORRECTIONS = [
    # Escalate issues and risks also produces an Issue Report (SME correction).
    ("activity::Escalate issues and risks", "product::Issue Report", "O", "indicative"),
]

_existing = {(r["from"], r["to"], r["code"]) for r in relationships}
for fk, tk, code, conf in CORRECTIONS:
    if fk not in entities or tk not in entities:
        raise SystemExit(f"correction references a missing entity: {fk} -> {tk}")
    if (fk, tk, code) not in _existing:
        relationships.append({"from": fk, "to": tk, "code": code, "confidence": conf})

data = {
    "framework": {
        "key": "prince2-7",
        "name": "PRINCE2 7",
        "edition": "7th edition (PeopleCert, 2023)",
        "description": (
            "PRINCE2 7 method map: 7 processes and their activities cross-referenced "
            "to management team roles, practices, management approaches and management "
            "products. Role/practice/approach codes: C=Responsible, P=Participates, "
            "N=Assists. Product codes: I=Input, O=Output, U=Update, A=Authorise. "
            "Activity rows and their codes are a best-effort reconstruction from public "
            "sources (prince2.wiki, CC-BY 4.0) and should be SME-verified before formal use."
        ),
        "sort_order": 1,
        # Framework definition: entity types (with colour, kind and layout zone),
        # relationship codes, and lifecycle lanes/phases. This is what makes the
        # app framework-agnostic — a new framework (e.g. MSP) supplies its own.
        # kind: container (owns children via parent_id) · hub (carries the coded
        # relationships) · node (a relationship target). zone: Matrix placement.
        "config": {
            "types": [
                {"key": "process", "label": "Processes", "color": "#0B2545", "kind": "container", "zone": "top", "order": 1},
                {"key": "activity", "label": "Activities", "color": "#3D5A80", "kind": "hub", "zone": "center", "order": 2},
                {"key": "role", "label": "Management Team Roles", "color": "#C9A227", "kind": "node", "code_group": "role", "zone": "left", "order": 3},
                {"key": "practice", "label": "Practices", "color": "#2E7D5B", "kind": "node", "code_group": "role", "zone": "right", "order": 4},
                {"key": "approach", "label": "Management Approaches", "color": "#8E5BE0", "kind": "node", "code_group": "role", "zone": "bottom", "order": 5},
                {"key": "product", "label": "Products", "color": "#C0392B", "kind": "node", "code_group": "product", "zone": "below", "order": 6},
            ],
            "codes": {
                "role": {"C": "Responsible", "P": "Participates", "N": "Assists"},
                "product": {"I": "Input", "O": "Output", "U": "Update", "A": "Authorise"},
            },
            "lanes": [
                {"key": "directing", "label": "Directing (Project Board)"},
                {"key": "managing", "label": "Managing (Project Manager)"},
                {"key": "delivering", "label": "Delivering (Team)"},
            ],
            # `column`/`header` mark phases that get a Timeline column header
            # (and its short label). Phases without them (stage-boundary, throughout)
            # don't head a time column.
            "phases": [
                {"key": "pre-project", "label": "Pre-project", "column": True, "header": "Pre-project"},
                {"key": "initiation", "label": "Initiation stage", "column": True, "header": "Initiation"},
                {"key": "delivery", "label": "Delivery stage(s)", "column": True, "header": "Delivery ⟳"},
                {"key": "stage-boundary", "label": "Stage boundary"},
                {"key": "final", "label": "Final delivery stage", "column": True, "header": "Final"},
                {"key": "throughout", "label": "Throughout"},
            ],
            # Worked-example documents (the "Helios" sample project). Keyed by the
            # exact product entity name; the value is the path, relative to the SPA
            # base, of a bundled PDF (frontend/public/examples/prince2/*.pdf). The
            # detail panel shows a "View worked example" button when a selected
            # product has an entry here. Sourced from the PMO Template Library; the
            # PRINCE2-7 "Product Register" uses the Product Status Account example.
            "examples": {
                "Business Case": "examples/prince2/business-case.pdf",
                "Plan": "examples/prince2/plan.pdf",
                "Product Description": "examples/prince2/product-description.pdf",
                "Project Brief": "examples/prince2/project-brief.pdf",
                "Project Initiation Documentation": "examples/prince2/project-initiation-documentation.pdf",
                "Project Product Description": "examples/prince2/project-product-description.pdf",
                "Work Package": "examples/prince2/work-package.pdf",
                "Daily Log": "examples/prince2/daily-log.pdf",
                "Issue Register": "examples/prince2/issue-register.pdf",
                "Lessons Log": "examples/prince2/lessons-log.pdf",
                "Product Register": "examples/prince2/product-register.pdf",
                "Quality Register": "examples/prince2/quality-register.pdf",
                "Risk Register": "examples/prince2/risk-register.pdf",
                "Checkpoint Report": "examples/prince2/checkpoint-report.pdf",
                "End Project Report": "examples/prince2/end-project-report.pdf",
                "End Stage Report": "examples/prince2/end-stage-report.pdf",
                "Exception Report": "examples/prince2/exception-report.pdf",
                "Highlight Report": "examples/prince2/highlight-report.pdf",
                "Issue Report": "examples/prince2/issue-report.pdf",
                "Lessons Report": "examples/prince2/lessons-report.pdf",
            },
        },
    },
    "entities": list(entities.values()),
    "relationships": relationships,
}

os.makedirs(os.path.dirname(OUT), exist_ok=True)
with open(OUT, "w", encoding="utf-8") as f:
    json.dump(data, f, indent=2, ensure_ascii=False)

# summary
from collections import Counter
c = Counter(e["type"] for e in entities.values())
print("entities:", dict(c), "total", len(entities))
print("relationships:", len(relationships))
print("rel code hist:", dict(Counter(r["code"] for r in relationships)))
print("wrote", OUT)
