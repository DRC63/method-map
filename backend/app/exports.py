"""CSV / XLSX / PDF export builders for client deliverables.

CSV and XLSX produce a flat cross-reference table (one row per relationship),
optionally narrowed to a single focus entity so the user can export exactly the
slice they're looking at. PDF produces a branded one-entity relationship summary.
"""
import csv
import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.enums import TA_LEFT
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)

from . import crud, models
from .enums import CODE_LABELS

NAVY = colors.HexColor("#0B2545")
ACCENT = colors.HexColor("#C9A227")
LIGHT = colors.HexColor("#EBEDF2")

HEADERS = ["Process", "Activity", "Relationship", "Target type", "Target", "Code", "Meaning", "Confidence"]


def _relationship_rows(
    db: models.Relationship, framework: models.Framework, entities: dict
) -> list[list[str]]:
    rels = crud.list_relationships(db, framework.id)
    rows = []
    for rel in rels:
        src = entities.get(rel.from_entity_id)
        tgt = entities.get(rel.to_entity_id)
        if not src or not tgt:
            continue
        process = src.parent.name if src.parent else ""
        rows.append(
            [
                process,
                src.name,
                "uses / touches",
                tgt.type,
                tgt.name,
                rel.code,
                CODE_LABELS.get(rel.code, rel.code),
                rel.confidence,
            ]
        )
    rows.sort(key=lambda r: (r[0], r[1], r[3], r[4]))
    return rows


def _filtered_rows(db, framework, focus_entity_id=None) -> list[list[str]]:
    entities = {e.id: e for e in crud.list_entities(db, framework.id)}
    rows = _relationship_rows(db, framework, entities)
    if focus_entity_id:
        focus = entities.get(focus_entity_id)
        if focus:
            # keep rows where the focus entity is either the activity or the target
            rows = [
                r
                for r in rows
                if r[1] == focus.name or r[4] == focus.name
            ]
    return rows


def export_csv(db, framework, focus_entity_id=None) -> bytes:
    rows = _filtered_rows(db, framework, focus_entity_id)
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow(HEADERS)
    writer.writerows(rows)
    return buf.getvalue().encode("utf-8-sig")


def export_xlsx(db, framework, focus_entity_id=None) -> bytes:
    rows = _filtered_rows(db, framework, focus_entity_id)
    wb = Workbook()
    ws = wb.active
    ws.title = framework.key[:31]

    header_fill = PatternFill("solid", fgColor="0B2545")
    header_font = Font(color="FFFFFF", bold=True)
    for col, header in enumerate(HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left")
    for r, row in enumerate(rows, start=2):
        for c, value in enumerate(row, start=1):
            ws.cell(row=r, column=c, value=value)

    widths = [26, 40, 16, 14, 34, 8, 16, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w
    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def entity_report_pdf(db, framework, entity: models.Entity) -> bytes:
    detail_rels = crud.relationships_for_entity(db, entity.id)
    entities = {e.id: e for e in crud.list_entities(db, framework.id)}

    buf = io.BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        topMargin=20 * mm,
        bottomMargin=18 * mm,
        leftMargin=18 * mm,
        rightMargin=18 * mm,
        title=f"{entity.name} - {framework.name} Method Map",
    )
    styles = getSampleStyleSheet()
    h1 = ParagraphStyle(
        "H1", parent=styles["Heading1"], textColor=NAVY, fontSize=18, spaceAfter=2
    )
    sub = ParagraphStyle(
        "Sub", parent=styles["Normal"], textColor=colors.HexColor("#5B6675"),
        fontSize=10, spaceAfter=12,
    )
    label = ParagraphStyle(
        "Label", parent=styles["Heading2"], textColor=NAVY, fontSize=12,
        spaceBefore=10, spaceAfter=4,
    )
    small = ParagraphStyle(
        "Small", parent=styles["Normal"], fontSize=8,
        textColor=colors.HexColor("#5B6675"), alignment=TA_LEFT, spaceBefore=14,
    )

    story = []
    story.append(Paragraph("P3MAI Method Map", sub))
    story.append(Paragraph(entity.name, h1))
    meta = f"{entity.type.title()}"
    if entity.code:
        meta += f" &middot; {entity.code}"
    meta += f" &middot; {framework.name} ({entity.confidence})"
    story.append(Paragraph(meta, sub))
    if entity.description:
        story.append(Paragraph(entity.description, styles["Normal"]))
        story.append(Spacer(1, 6))

    # group related entities
    outgoing = []
    incoming = {}
    for rel in detail_rels:
        if rel.from_entity_id == entity.id:
            other = entities.get(rel.to_entity_id)
            if other:
                outgoing.append((other, rel))
        else:
            other = entities.get(rel.from_entity_id)
            if other:
                proc = other.parent.name if other.parent else "(no process)"
                incoming.setdefault(proc, []).append((other, rel))

    def make_table(pairs):
        data = [["Type", "Name", "Code", "Meaning", "Confidence"]]
        for other, rel in sorted(pairs, key=lambda p: (p[0].type, p[0].name)):
            data.append(
                [
                    other.type.title(),
                    other.name,
                    rel.code,
                    CODE_LABELS.get(rel.code, rel.code),
                    rel.confidence,
                ]
            )
        t = Table(data, colWidths=[24 * mm, 62 * mm, 14 * mm, 32 * mm, 24 * mm])
        t.setStyle(
            TableStyle(
                [
                    ("BACKGROUND", (0, 0), (-1, 0), NAVY),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 8.5),
                    ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, LIGHT]),
                    ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#D5D9E0")),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("TOPPADDING", (0, 0), (-1, -1), 4),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
                ]
            )
        )
        return t

    if outgoing:
        story.append(Paragraph("This activity uses / produces", label))
        story.append(make_table(outgoing))
    if incoming:
        story.append(
            Paragraph("Referenced by activities (grouped by process)", label)
        )
        for proc, pairs in sorted(incoming.items()):
            story.append(Paragraph(proc, styles["Heading3"]))
            story.append(make_table(pairs))
            story.append(Spacer(1, 4))
    if not outgoing and not incoming:
        story.append(Paragraph("No relationships recorded yet.", styles["Normal"]))

    story.append(
        Paragraph(
            "Generated by P3MAI Method Map. Relationship codes: "
            "C=Responsible, P=Participates, N=Assists (roles / practices / "
            "approaches); I=Input, O=Output, U=Update, A=Authorise (products). "
            "Activity-level detail is indicative and should be SME-verified "
            "against the licensed manual before formal use.",
            small,
        )
    )
    doc.build(story)
    return buf.getvalue()
